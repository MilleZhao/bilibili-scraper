'''
弹幕去重 + 按重复次数排序 + 导出 Excel (v2: 保留? 剔除块状装饰符)
运行: python export_danmaku.py
'''
import json, re, os
from pathlib import Path
from datetime import datetime, timezone
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

SRC = Path(r'C:\Users\LENOVO\Documents\爬虫\outputs\danmaku_BV11mFLziEyP.json')
raw = json.loads(SRC.read_text(encoding='utf-8'))

for d in raw:
    d['text'] = ILLEGAL_CHARACTERS_RE.sub('', d.get('text', ''))

# ============================================================
# 清洗规则 (v2)
# ============================================================
# 块状装饰符: Unicode Block Elements U+2580-259F + 常见全角装饰
BLOCK_SYMBOLS = (
    r'\u2580-\u259f'              # ▀▁▂▃▄▅▆▇█▉▊▋▌▍▎▏▐░▒▓▔▕▖▗▘▙▚▛▜▝▞▟
    r'\u25a0-\u25ff'              # ■□▢▣▤▥▦▧▨▩▪▫▬▭▮▯▰▱▲△▴▵▶▸▹►▻▼▽▾▿
    r'\u2600-\u26ff'              # ☀☁★☆☇...
    r'\u2700-\u27bf'              # ✀✁✂✅...
    r'\u2b00-\u2bff'              # ⬀⬁...
)

PURE_BLOCK_RE = re.compile(
    r'^[\d\s\.\,\!\?\@\$\%\^\&\*\(\)\+\=~\-—–_/\\|:;'
    + BLOCK_SYMBOLS
    + r']+$'
)

# 可保留的单字情绪符号 (仅?/？/！不触发单字剔除)
EMOTION_SINGLE = {'?', '？', '!', '！'}

def is_valid_danmaku(d):
    text = (d.get('text') or '').strip()
    if not text:
        return False, 'empty'

    # 单字剔除, 但保留 ?/？/!
    if len(text) == 1:
        if text in EMOTION_SINGLE:
            return True, None
        return False, 'single'

    # 块状装饰符剔除
    if PURE_BLOCK_RE.match(text):
        return False, 'block'

    return True, None

clean = []
stats_empty  = 0
stats_single = 0
stats_block  = 0

for d in raw:
    ok, reason = is_valid_danmaku(d)
    if ok:
        clean.append(d)
    elif reason == 'empty':
        stats_empty += 1
    elif reason == 'single':
        stats_single += 1
    elif reason == 'block':
        stats_block += 1

print('=== Danmaku Cleaning (v2) ===')
print(f'  Raw:          {len(raw):>8,}')
print(f'  - Empty:       {stats_empty:>8,}')
print(f'  - Single-char: {stats_single:>8,}  (except ?/？/!)')
print(f'  - Block decor: {stats_block:>8,}  (\u2580-\u259f etc)')
print(f'  = After clean: {len(clean):>8,}')

# ============================================================
# 去重
# ============================================================
groups = {}
for d in clean:
    text = d['text']
    if text not in groups:
        groups[text] = {
            'text': text, 'count': 0,
            'first_at': d.get('appear_at', 0),
            'first_send': d.get('send_time', ''),
            'modes': Counter(), 'colors': Counter(),
            'pools': Counter(), 'dates': Counter(),
        }
    g = groups[text]
    g['count'] += 1
    g['modes'][d.get('mode', 0)] += 1
    g['colors'][d.get('color', 'ffffff')] += 1
    g['pools'][d.get('pool', 0)] += 1
    g['dates'][d.get('date', '')] += 1
    at = d.get('appear_at', 0)
    if at < g['first_at']:
        g['first_at'] = at
    st = d.get('send_time', '')
    if st and (not g['first_send'] or st < g['first_send']):
        g['first_send'] = st

dedup = sorted(groups.values(), key=lambda g: g['count'], reverse=True)

print(f'  After dedup:   {len(dedup):>8,}  unique texts')
print(f'  Max repeat:    {dedup[0]["count"]}x — {dedup[0]["text"][:40]}')

dm_repeat_total = sum(g['count'] for g in dedup)
dm_repeat_max   = dedup[0]['count'] if dedup else 0
dedup_rate      = len(dedup) / max(len(raw), 1) * 100

print(f'  Total repeats: {dm_repeat_total:,}')
print(f'  Dedup ratio:   {dedup_rate:.1f}%')

# ============================================================
# 删除旧 Excel
# ============================================================
old_files = [
    SRC.parent / 'danmaku_sorted.xlsx',
    SRC.parent / 'danmaku_dedup.xlsx',
]
for f in old_files:
    if f.exists():
        f.unlink()
        print(f'  Deleted old: {f.name}')

# ============================================================
# Excel 导出
# ============================================================
OUT = SRC.parent / 'danmaku_dedup.xlsx'
wb = Workbook()
ws = wb.active
ws.title = 'B站弹幕(去重)'

pink   = PatternFill(start_color='FB7299', end_color='FB7299', fill_type='solid')
gold   = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
hot    = PatternFill(start_color='FDE8E8', end_color='FDE8E8', fill_type='solid')
header_font = Font(name='\u5fae\u8f6f\u96c5\u9ed1', size=11, bold=True, color='FFFFFF')
cell_font   = Font(name='\u5fae\u8f6f\u96c5\u9ed1', size=10)
thin = Border(
    left=Side(style='thin', color='E0E0E0'), right=Side(style='thin', color='E0E0E0'),
    top=Side(style='thin', color='E0E0E0'), bottom=Side(style='thin', color='E0E0E0'),
)

headers = ['\u6392\u540d', '\u5f39\u5e55\u5185\u5bb9', '\u91cd\u590d\u6b21\u6570', '\u9996\u6b21\u51fa\u73b0(s)', '\u6700\u65e9\u53d1\u9001', '\u4e3b\u8981\u7c7b\u578b', '\u4e3b\u8981\u989c\u8272', '\u4e3b\u8981\u5f39\u5e55\u6c60', '\u51fa\u73b0\u5929\u6570']
widths  = [6, 50, 10, 12, 18, 10, 10, 10, 10]
type_map = {1: '\u6eda\u52a8', 4: '\u5e95\u90e8', 5: '\u9876\u90e8'}
pool_map = {0: '\u666e\u901a', 1: '\u5b57\u5e55', 2: '\u7279\u6b8a'}

for ci, (h, w) in enumerate(zip(headers, widths), 1):
    c = ws.cell(row=1, column=ci, value=h)
    c.fill, c.font, c.border = pink, header_font, thin
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.column_dimensions[c.column_letter].width = w

for rank, g in enumerate(dedup, 1):
    r = rank + 1
    ts = g['first_send']
    if ts and str(ts).isdigit():
        try:
            ts_dt = datetime.fromtimestamp(int(ts), tz=timezone.utc).astimezone()
            ts_str = ts_dt.strftime('%Y-%m-%d %H:%M')
        except:
            ts_str = str(ts)
    else:
        ts_str = str(ts) if ts else ''

    top_mode  = g['modes'].most_common(1)[0] if g['modes'] else (1, 0)
    top_color = g['colors'].most_common(1)[0] if g['colors'] else ('ffffff', 0)
    top_pool  = g['pools'].most_common(1)[0] if g['pools'] else (0, 0)

    vals = [
        rank, g['text'], g['count'],
        round(g['first_at'], 1), ts_str,
        type_map.get(top_mode[0], '\u6eda\u52a8'),
        '#' + str(top_color[0]),
        pool_map.get(top_pool[0], '\u666e\u901a'),
        len(g['dates']),
    ]

    for ci, val in enumerate(vals, 1):
        c = ws.cell(row=r, column=ci, value=val)
        c.font, c.border = cell_font, thin
        if ci == 2:
            c.alignment = Alignment(vertical='center', wrap_text=True)
        else:
            c.alignment = Alignment(horizontal='center', vertical='center')

    if g['count'] >= 50:
        for ci in range(1, 10):
            ws.cell(row=r, column=ci).fill = hot
    elif g['count'] >= 20:
        for ci in range(1, 10):
            ws.cell(row=r, column=ci).fill = gold

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:I{len(dedup)+1}'
wb.save(str(OUT))

# ============================================================
# 保存 JSON 供报告使用
# ============================================================
dedup_json = SRC.parent / 'danmaku_dedup.json'
output = []
for g in dedup:
    output.append({
        'text': g['text'], 'count': g['count'],
        'first_at': g['first_at'], 'first_send': g['first_send'],
        'top_mode':  g['modes'].most_common(1)[0][0] if g['modes'] else 1,
        'top_color': g['colors'].most_common(1)[0][0] if g['colors'] else 'ffffff',
        'top_pool':  g['pools'].most_common(1)[0][0] if g['pools'] else 0,
        'date_count': len(g['dates']),
    })
dedup_json.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding='utf-8')

# 导出清洗阶段统计供报告使用
stage_stats = {
    'raw':    len(raw),
    'empty':  stats_empty,
    'single': stats_single,
    'block':  stats_block,
    'clean':  len(clean),
    'dedup':  len(dedup),
    'max_repeat': dm_repeat_max,
    'total_repeat': dm_repeat_total,
    'dedup_rate': round(dedup_rate, 1),
}
stage_json = SRC.parent / 'danmaku_stage_stats.json'
stage_json.write_text(json.dumps(stage_stats, ensure_ascii=False, indent=2), encoding='utf-8')

print(f'\nDone!')
print(f'  JSON:  {dedup_json}')
print(f'  Excel: {OUT} ({len(dedup):,} rows)')
