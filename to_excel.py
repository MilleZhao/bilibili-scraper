"""一键生成 Excel：读取 + 去无效 + 排序 + 导出"""
import json, re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

SRC = Path(__file__).parent / "outputs" / "comments_all.json"
OUT = SRC.parent / "comments_sorted.xlsx"

# ---- 1. 读取 + 清洗非法字符 ----
raw = json.loads(SRC.read_text(encoding="utf-8"))
for item in raw:
    if item.get("content"):
        item["content"] = ILLEGAL_CHARACTERS_RE.sub("", item["content"])
    if item.get("username"):
        item["username"] = ILLEGAL_CHARACTERS_RE.sub("", item["username"])

# ---- 2. 判断函数 ----
def is_emoji(ch):
    """判断单个字符是否为 emoji"""
    cp = ord(ch)
    return (
        (0x1F300 <= cp <= 0x1FAFF) or
        (0x2600 <= cp <= 0x27BF) or
        cp == 0x2764 or
        (0x1F900 <= cp <= 0x1F9FF) or
        (0x1FA00 <= cp <= 0x1FA6F) or
        (0x200D <= cp <= 0x200D) or         # ZWJ
        (0xFE0F <= cp <= 0xFE0F) or         # variant selector
        (0x2702 <= cp <= 0x27B0) or
        (0x1F600 <= cp <= 0x1F64F) or       # emoticons
        (0x1F680 <= cp <= 0x1F6FF) or       # transport
        (0x1F1E0 <= cp <= 0x1F1FF)          # flags
    )

def has_text(text):
    """去掉 emoji/颜文字后是否有实质性内容"""
    # 去 emoji
    cleaned = "".join(ch for ch in text if not is_emoji(ch))
    # 去颜文字括号
    cleaned = re.sub(r"[\(（]\s*[^\w\s]{2,}\s*[\)）]", "", cleaned)
    chinese = len(re.findall(r"[\u4e00-\u9fff]", cleaned))
    english = len(re.findall(r"[a-zA-Z]{3,}", cleaned))
    return chinese >= 2 or english >= 1

def is_ad(text):
    URL_RE    = re.compile(r"https?://\S+|www\.\S+", re.I)
    PHONE_RE  = re.compile(r"1[3-9]\d{9}|\d{3}[-\s]?\d{4}[-\s]?\d{4}")
    WECHAT_RE = re.compile(r"[加➕﹢＋]?\s*(微|v信|威信|VX|vx|WX|wx)[:：]?\s*\w{5,}", re.I)
    QQ_RE     = re.compile(r"[加➕﹢＋]?\s*(qq|QQ|扣扣|Q|q)[:：]?\s*\d{5,}", re.I)
    SCAN_RE   = re.compile(r"扫码|二维码|QR码")
    PROMO_RE  = re.compile(r"免费领取|免费送|低价|优惠券|薅羊毛|返利|刷单|兼职|日赚|代练|代打")
    return bool(URL_RE.search(text) or PHONE_RE.search(text) or WECHAT_RE.search(text)
             or QQ_RE.search(text) or SCAN_RE.search(text) or PROMO_RE.search(text))

def is_valid(c):
    text = (c.get("content") or "").strip()
    if not text:
        return False
    if not has_text(text):       # 纯表情/颜文字
        return False
    if is_ad(text):              # 广告
        return False
    if len(text) <= 2:           # 过短
        return False
    if text in ("好","加油","支持","6","厉害","牛","顶","打卡","来了","第一",
                "棒","不错","nb","NB","哈哈","哈哈哈","可以的","好家伙","三连"):
        return False
    if re.fullmatch(r"[\d\s,，]+楼", text):      # 抢楼
        return False
    if len(set(text)) <= 3 and len(text) >= 3:   # 重复单字
        return False
    if re.fullmatch(r"[\d\s\.\,\!\?\#\@\$\%\^\&\*\(\)\+\=~\-—–_/\\|:;]+", text):  # 纯符号
        return False
    return True

valid = [c for c in raw if is_valid(c)]
valid.sort(key=lambda c: c.get("likes", 0), reverse=True)

# 统计
ad_count    = sum(1 for c in raw if is_ad(c.get("content","")))
emoji_count = sum(1 for c in raw if not has_text(c.get("content","")))
other_count = len(raw) - len(valid) - ad_count - emoji_count
print(f"原始: {len(raw):,}")
print(f"  ├─ 纯表情/颜文字: {emoji_count:,}")
print(f"  ├─ 广告:          {ad_count:,}")
print(f"  ├─ 其他无效:      {other_count:,}")
print(f"  └─ 有效:          {len(valid):,}")

# ---- 3. 生成 Excel ----
wb = Workbook()
ws = wb.active
ws.title = "B站评论"

header_fill = PatternFill(start_color="FB7299", end_color="FB7299", fill_type="solid")
header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
cell_font   = Font(name="微软雅黑", size=10)
thin_border = Border(
    left=Side(style="thin", color="E0E0E0"),
    right=Side(style="thin", color="E0E0E0"),
    top=Side(style="thin", color="E0E0E0"),
    bottom=Side(style="thin", color="E0E0E0"),
)
highlight = PatternFill(start_color="FFF0F5", end_color="FFF0F5", fill_type="solid")

headers = ["排名","类型","用户名","用户ID","评论内容","发布时间","点赞数","回复数","评论ID"]
widths  = [6, 6, 16, 12, 60, 20, 10, 8, 14]

for ci, (h, w) in enumerate(zip(headers, widths), 1):
    c = ws.cell(row=1, column=ci, value=h)
    c.fill, c.font, c.border = header_fill, header_font, thin_border
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[c.column_letter].width = w

for rank, row in enumerate(valid, 1):
    r = rank + 1
    ts = row.get("ctime", 0)
    ts_str = datetime.fromtimestamp(ts, tz=timezone.utc).astimezone().strftime("%Y-%m-%d %H:%M") if ts else ""

    vals = [
        rank, "子回复" if row["is_sub"] else "主评论",
        row.get("username",""), row.get("user_mid",""),
        row.get("content",""), ts_str,
        row.get("likes",0), row.get("reply_count",0), row.get("rpid",""),
    ]

    for ci, val in enumerate(vals, 1):
        c = ws.cell(row=r, column=ci, value=val)
        c.font, c.border = cell_font, thin_border
        if ci == 5:
            c.alignment = Alignment(vertical="center", wrap_text=True)
        elif ci in (1,2,7,8):
            c.alignment = Alignment(horizontal="center", vertical="center")

    if row.get("likes", 0) >= 100:
        for ci in range(1, 10):
            ws.cell(row=r, column=ci).fill = highlight

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:I{len(valid)+1}"

wb.save(str(OUT))

# ---- 4. 预览 ----
print(f"\nTop 10 (按点赞):")
for i, c in enumerate(valid[:10], 1):
    tag = "[子]" if c["is_sub"] else "[主]"
    print(f"  {i:>2}. {tag} 👍{c['likes']:>6}  {c['username']}: {c['content'][:70]}")

print(f"\n✅ Excel: {OUT}")
